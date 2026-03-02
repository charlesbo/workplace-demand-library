"""Excel exporter for workplace demands."""

import json
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import func

from src.storage.database import get_session
from src.storage.models import Demand
from src.utils.config import get_settings
from src.utils.logger import get_logger

logger = get_logger(__name__)

HEADERS = ["ID", "标题", "分类", "子分类", "频次", "重要性", "趋势", "状态", "首次发现", "最后发现", "标签"]

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
CATEGORY_FILL = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")


def _demand_to_row(d: Demand) -> list:
    """Convert a Demand model instance to a spreadsheet row."""
    tags = ""
    if d.tags:
        try:
            tags = ", ".join(json.loads(d.tags))
        except (json.JSONDecodeError, TypeError):
            tags = d.tags
    return [
        d.id,
        d.title,
        d.category or "",
        d.subcategory or "",
        d.frequency,
        round(d.importance_score, 2),
        d.trend,
        d.status,
        d.first_seen.strftime("%Y-%m-%d") if d.first_seen else "",
        d.last_seen.strftime("%Y-%m-%d") if d.last_seen else "",
        tags,
    ]


def _style_header(ws, num_cols: int) -> None:
    """Apply bold + color fill styling to the first row."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _auto_width(ws) -> None:
    """Auto-adjust column widths based on content."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                cell_len = len(str(cell.value)) if cell.value is not None else 0
                # CJK characters are roughly double-width
                cjk_count = sum(1 for ch in str(cell.value or "") if ord(ch) > 0x4E00)
                cell_len += cjk_count
                max_len = max(max_len, cell_len)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)


class ExcelExporter:
    """Export workplace demands to an Excel workbook with multiple sheets."""

    def __init__(self) -> None:
        self.settings = get_settings()

    def export(self, output_path: Optional[str] = None) -> str:
        """Export all demands to an Excel file.

        Args:
            output_path: Destination file path. Defaults to
                ``<export.output_dir>/demands_YYYYMMDD.xlsx``.

        Returns:
            The absolute path of the generated Excel file.
        """
        if output_path is None:
            output_dir = Path(self.settings.get("export.output_dir", "output"))
            output_dir.mkdir(parents=True, exist_ok=True)
            output_path = str(output_dir / f"demands_{datetime.now():%Y%m%d}.xlsx")
        else:
            Path(output_path).parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        demands, categories = self._load_data()

        # Sheet 1 — 需求总表
        ws_all = wb.active
        ws_all.title = "需求总表"
        ws_all.append(HEADERS)
        for d in demands:
            ws_all.append(_demand_to_row(d))
        _style_header(ws_all, len(HEADERS))
        _auto_width(ws_all)

        # Per-category sheets
        for cat, cat_demands in categories.items():
            ws = wb.create_sheet(title=cat[:31])  # sheet name max 31 chars
            ws.append(HEADERS)
            for d in cat_demands:
                ws.append(_demand_to_row(d))
            _style_header(ws, len(HEADERS))
            _auto_width(ws)

        # 统计 sheet
        self._build_stats_sheet(wb, demands, categories)

        wb.save(output_path)
        logger.info(f"Excel exported: {output_path}")
        return output_path

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------

    def _load_data(self) -> tuple[list[Demand], dict[str, list[Demand]]]:
        """Load demands from the database, grouped by category."""
        with get_session() as session:
            demands: list[Demand] = (
                session.query(Demand).order_by(Demand.frequency.desc()).all()
            )
            session.expunge_all()

        categories: dict[str, list[Demand]] = {}
        for d in demands:
            cat = d.category or "未分类"
            categories.setdefault(cat, []).append(d)
        return demands, categories

    def _build_stats_sheet(
        self,
        wb: Workbook,
        demands: list[Demand],
        categories: dict[str, list[Demand]],
    ) -> None:
        """Create the 统计 summary sheet."""
        ws = wb.create_sheet(title="统计")
        total = len(demands)

        # Category counts
        stat_headers = ["分类", "数量", "占比", "平均频次", "平均重要性"]
        ws.append(stat_headers)
        for cat, cat_demands in sorted(categories.items(), key=lambda x: -len(x[1])):
            count = len(cat_demands)
            pct = f"{count / total * 100:.1f}%" if total else "0%"
            avg_freq = sum(d.frequency for d in cat_demands) / count if count else 0
            avg_imp = (
                sum(d.importance_score for d in cat_demands) / count if count else 0
            )
            ws.append([cat, count, pct, round(avg_freq, 1), round(avg_imp, 2)])

        # Totals row
        ws.append(["合计", total, "100%", "", ""])

        _style_header(ws, len(stat_headers))
        _auto_width(ws)
